"""
Unit tests for target writers: CSV, JSON, Excel output verification.
"""
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import csv
import json
import tempfile
import pytest
from services.mapping.writers import CsvWriter, JsonWriter, ExcelWriter, get_writer


SAMPLE_ROWS = [
    {"account": "VGD-500", "valuationDate": "2026-03-09", "totalNetAssets": 1234567890.12, "navPerShare": 27.03},
    {"account": "FID-CONTRA", "valuationDate": "2026-03-09", "totalNetAssets": 2345678901.23, "navPerShare": 29.73},
]
FIELD_NAMES = ["account", "valuationDate", "totalNetAssets", "navPerShare"]


class TestCsvWriter:
    def test_write_csv(self):
        writer = CsvWriter()
        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False, mode="w") as f:
            path = f.name

        try:
            result = writer.write(SAMPLE_ROWS, path, {"delimiter": ","}, FIELD_NAMES)
            assert result == path
            with open(path, "r") as f:
                reader = csv.DictReader(f)
                rows = list(reader)
            assert len(rows) == 2
            assert rows[0]["account"] == "VGD-500"
            assert rows[1]["navPerShare"] == "29.73"
        finally:
            os.unlink(path)

    def test_write_empty(self):
        writer = CsvWriter()
        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as f:
            path = f.name
        try:
            writer.write([], path, {})
            assert os.path.getsize(path) == 0
        finally:
            os.unlink(path)


class TestJsonWriter:
    def test_write_json_array(self):
        writer = JsonWriter()
        with tempfile.NamedTemporaryFile(suffix=".json", delete=False) as f:
            path = f.name
        try:
            writer.write(SAMPLE_ROWS, path, {"arrayWrapper": True, "prettyPrint": False})
            with open(path, "r") as f:
                data = json.load(f)
            assert isinstance(data, list)
            assert len(data) == 2
            assert data[0]["account"] == "VGD-500"
        finally:
            os.unlink(path)

    def test_write_json_pretty(self):
        writer = JsonWriter()
        with tempfile.NamedTemporaryFile(suffix=".json", delete=False) as f:
            path = f.name
        try:
            writer.write(SAMPLE_ROWS, path, {"arrayWrapper": True, "prettyPrint": True})
            with open(path, "r") as f:
                content = f.read()
            assert "\n" in content  # Pretty printed
            data = json.loads(content)
            assert len(data) == 2
        finally:
            os.unlink(path)


class TestExcelWriter:
    def test_write_excel(self):
        writer = ExcelWriter()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            writer.write(SAMPLE_ROWS, path, {"sheetName": "Output"}, FIELD_NAMES)
            from openpyxl import load_workbook
            wb = load_workbook(path)
            ws = wb["Output"]
            headers = [cell.value for cell in ws[1]]
            assert headers == FIELD_NAMES
            assert ws.cell(2, 1).value == "VGD-500"
            assert ws.max_row == 3  # header + 2 data rows
        finally:
            os.unlink(path)


class TestWriterFactory:
    def test_get_csv_writer(self):
        assert isinstance(get_writer("CSV"), CsvWriter)

    def test_get_json_writer(self):
        assert isinstance(get_writer("JSON"), JsonWriter)

    def test_get_excel_writer(self):
        assert isinstance(get_writer("EXCEL"), ExcelWriter)

    def test_unsupported(self):
        with pytest.raises(ValueError):
            get_writer("XML")
