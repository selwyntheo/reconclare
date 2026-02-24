"""Excel/PDF export generation endpoints."""
from datetime import datetime
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
from db.mongodb import get_async_db, COLLECTIONS

router = APIRouter(prefix="/api", tags=["export"])


@router.post("/export/excel")
async def export_to_excel(body: dict):
    """Generate Excel export for a grid view."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        return {"error": "openpyxl is not installed. Run: pip install openpyxl"}

    view_type = body.get("viewType", "generic")
    event_id = body.get("eventId", "")
    filters = body.get("filters", {})
    exported_by = body.get("exportedBy", "System")

    db = get_async_db()

    wb = Workbook()
    ws = wb.active
    ws.title = view_type.replace("-", " ").title()

    # Metadata header
    ws.append([f"Event: {event_id} | Exported: {datetime.utcnow().isoformat()} | By: {exported_by}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws["A1"].font = Font(bold=True, size=11)
    ws.append([])  # Blank row

    # RAG fill colors
    green_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    amber_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    red_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

    # Fetch data based on view type
    data: list = []
    headers: list = []

    if view_type == "client-scorecard":
        headers = ["Fund Account", "Fund Name", "BNY Net Assets", "Incumbent Net Assets",
                    "Net Assets Diff", "BP Diff", "RAG", "Adjusted Diff", "Adjusted BP", "Adjusted RAG"]
        event = await db[COLLECTIONS["events"]].find_one({"eventId": event_id}, {"_id": 0})
        if event:
            for fund in event.get("funds", []):
                data.append([fund["account"], fund.get("fundName", ""), 0, 0, 0, 0, "", 0, 0, ""])

    elif view_type == "nav-fund-level":
        headers = ["Valuation Date", "Account", "Account Name", "Incumbent TNA", "BNY TNA",
                    "TNA Difference", "BP Difference", "RAG"]

    else:
        headers = ["Column1", "Column2", "Column3"]

    # Task 20.3 — Use write-only mode for large datasets (>1000 rows)
    if len(data) > 1000:
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title=view_type.replace("-", " ").title())
        # Write-only mode does not support merge_cells or cell styling by
        # reference, so the metadata header is skipped for large datasets.
        ws.append(headers)
        for row in data:
            ws.append(row)
    else:
        # Write headers with styling (standard mode)
        ws.append(headers)
        header_row = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

        # Write data
        for row in data:
            ws.append(row)

        # Auto-width columns
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)

    # Write to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{view_type}_{event_id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
